import openpyxl.worksheet
import numpy as np


# Names of songs
SONGS: dict[str, int] = {
    "Long Ago and Far Away": 0,
    "Summertime": 1,
    "My Little Suede Shoes": 2
}

# Names of algorithms
ALGORITHMS: dict[str, int] = {
    "Note Retrieval": 0,
    "Token Factor Oracle": 1,
    "Token Markov": 2,
}


def get_responses() -> np.ndarray[int]:
    # For every pupil, for every song, for every algorithm, there are 4 responses
    responses: np.ndarray[int] = np.zeros((5, 3, 3, 4), dtype=int)
    workbook: openpyxl.Workbook = openpyxl.load_workbook("spreadsheet.xlsx")

    # For every pupil
    for pupil in range(5):
        # For every session they do
        for session in range(3):
            # For every performance they do within that session
            for performance in range(3):
                # Get Excel sheet, get starting row
                sheet = workbook[str(pupil + 1)]
                start_row: int = 8 * session + 27 * performance

                # Get performance info as text
                song_and_algorithm: str = str(sheet.cell(start_row + 2, 3).value)
                # Extract song and algorithm from text
                [song_str, algorithm_str] = song_and_algorithm.split(", ")
                song: int = SONGS[song_str]
                algorithm: int = ALGORITHMS[algorithm_str]

                # Set responses
                for i in range(4):
                    responses[pupil][song][algorithm][i] = int(sheet.cell(start_row + 3 + i, 6).value)

    return responses
