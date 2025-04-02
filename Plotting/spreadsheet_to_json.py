import openpyxl.worksheet
from dataclasses import dataclass
from recording import Recording
from pupil_info import *
import json

workbook: openpyxl.Workbook = openpyxl.load_workbook("spreadsheet.xlsx")

res = {
    "pupils": []
}

# For every pupil
for pupil in range(NUM_PUPILS):
    pupil_info = {
        "sessions": []
    }

    # Get Excel sheet, get starting row
    sheet = workbook[str(pupil + 1)]

    # For every session they do
    for session in range(NUM_SESSIONS):
        session_info = {
            "performance": []
        }

        # For every performance they do within that session
        for performance in range(NUM_PERFORMANCES):
            # Get start row in spreadsheet
            start_row: int = 8 * performance + 27 * session

            # Get performance info as text
            song_and_algorithm: str = str(sheet.cell(start_row + 2, 3).value)

            # Extract song and algorithm from text
            [song_str, algorithm_str] = song_and_algorithm.split(", ")
            song: int = SONG_IDS[song_str]
            algorithm: int = ALGORITHM_IDS[algorithm_str]

            # Get self-reported scores
            scores: list[int] = [int(sheet.cell(start_row + 3 + i, 6).value) for i in range(NUM_QUESTIONS)]

            performance_info = {
                "song": song,
                "algorithm": algorithm,
                "scores": scores,
            }

            session_info["performance"].append(performance_info)
        
        pupil_info["sessions"].append(session_info)
    
    res["pupils"].append(pupil_info)

with open("info.json", 'w') as f:
    res_json = json.dump(res, f, indent=2)
